#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
@author: PlasticMem
@time: 2021/8/29 13:29:52
@function:
"""

import os
from openpyxl import load_workbook


class XlsxWorkbook:
    """读取xlsx格式的Excel文件"""

    def __init__(self, filename: str = '', sheet_name: str = ''):
        self.workbook = None
        self.sheet = None

        if filename:
            self.open_workbook(filename)
        if self.workbook is not None and sheet_name:
            self.select_sheet(sheet_name)

    @staticmethod
    def get_row_value(row: tuple):
        """获取某一行每个单元格的数据，作为列表输出"""
        if not row:
            return None
        return [cell.value for cell in row]

    def open_workbook(self, filename):
        """打开workbook"""
        if not os.path.exists(filename):
            raise SystemError("Excel文件不存在")
        self.workbook = load_workbook(filename=filename)

    def select_sheet(self, sheet_name):
        """选择工作表"""
        if self.workbook is None:
            raise ValueError("尚未打开Excel文件")
        self.sheet = self.workbook[sheet_name]

    def read_sheet_header(self):
        """读取表头"""
        if self.sheet is None:
            raise ValueError("尚未选择Sheet")
        if self.sheet.max_row == 0:
            return None
        return self.get_row_value(self.sheet[1])

    def read_iter_sheet_data(self):
        """
        迭代读取每行表数据，不包含表头（即表的第一行数据）
        返回一个表数据的迭代器
        每次迭代返回的行需要通过"get_row_value"方法获取每个单元格的数据
        """
        if self.sheet is None:
            raise ValueError("尚未选择Sheet")
        if self.sheet.max_row < 2:
            return None
        iter_rows = self.sheet.iter_rows()
        next(iter_rows)
        return iter_rows

    def read_sheet_row(self, row_num: int):
        """读取表的某一行数据"""
        if self.sheet is None:
            raise ValueError("尚未选择Sheet")
        if row_num < 1:
            raise ValueError("行号不能小于1")
        if row_num > self.sheet.max_row:
            return None
        return self.get_row_value(self.sheet[row_num])

    def read_sheet_cell(self, row_num: int, header: str):
        """根据行号和表头名称读取某个单元格的数据"""
        headers = self.read_sheet_header()
        if headers is None:
            return None
        if header not in headers:
            return None

        row = self.read_sheet_row(row_num)
        if row is None:
            return None

        return row[headers.index(header)]

    def get_row_cell_value(self, row: tuple, header: str):
        """根据表头名称获取某一行对应的单元格的数据"""
        if not row:
            return None

        headers = self.read_sheet_header()
        if headers is None:
            return None
        if header not in headers:
            return None

        return row[headers.index(header)].value
